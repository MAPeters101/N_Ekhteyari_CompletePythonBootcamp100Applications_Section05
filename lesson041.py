import openpyxl

workbook = openpyxl.Workbook()
sheet = workbook.active

data = [['Name','Age','City'],
        ['Alice',25,'New York'],
        ['Bob',30,'San Francisco'],
        ['Charlie',22,'Los Angeles']]

for row in data:
    sheet.append(row)

workbook.save("new_file041.xlsx")

loaded_workbook = openpyxl.load_workbook("new_file041.xlsx")
loaded_sheet = loaded_workbook.active

for row in loaded_sheet.iter_rows(values_only=True):
    print(row)


